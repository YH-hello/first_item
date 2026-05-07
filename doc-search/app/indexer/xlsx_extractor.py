"""
Excel(XLSX) 텍스트 및 시트 정보 추출기
라이브러리: openpyxl
딥링크: win32com Excel.Application → Sheets.Activate()
"""
import logging
from typing import List

import openpyxl

from .base_extractor import BaseExtractor, PageChunk

logger = logging.getLogger(__name__)


class XLSXExtractor(BaseExtractor):
    """openpyxl을 이용한 시트별 텍스트 및 셀 주소 추출"""

    def extract(self) -> List[PageChunk]:
        chunks: List[PageChunk] = []
        try:
            wb = openpyxl.load_workbook(
                str(self.file_path), read_only=True, data_only=True
            )

            abs_path = str(self.file_path.resolve()).replace("\\", "\\\\")

            for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
                ws = wb[sheet_name]
                row_lines: List[str] = []

                for row in ws.iter_rows():
                    row_parts = []
                    for cell in row:
                        if cell.value is not None:
                            val = str(cell.value).strip()
                            if val:
                                row_parts.append(f"{cell.coordinate}:{val}")
                    if row_parts:
                        row_lines.append(" | ".join(row_parts))

                text = "\n".join(row_lines)
                if not text.strip():
                    continue

                # 딥링크: 시트 이름 + 셀 주소 (예: 'Sheet1'!A1)
                cell_ref = f"'{sheet_name}'!A1"
                page_label = f"'{sheet_name}' 시트"
                page_link_cmd = (
                    f"python -c \""
                    f"import win32com.client; "
                    f"xl=win32com.client.Dispatch('Excel.Application'); "
                    f"xl.Visible=True; "
                    f"wb=xl.Workbooks.Open(r'{abs_path}'); "
                    f"wb.Sheets('{sheet_name}').Activate()\""
                )

                chunks.append(
                    self._build_chunk(
                        text=text,
                        page_num=sheet_idx,
                        page_label=page_label,
                        page_link_cmd=page_link_cmd,
                        extra_meta={
                            "sheet_name": sheet_name,
                            "cell_ref": cell_ref,
                            "total_sheets": len(wb.sheetnames),
                        },
                    )
                )

            wb.close()

        except Exception as exc:
            logger.error("[XLSX] 추출 오류 %s: %s", self.file_path, exc, exc_info=True)

        return chunks
